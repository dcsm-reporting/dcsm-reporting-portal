#!/usr/bin/env python3
"""One-time seed of the Friends table from the "Baptisms (MLC)" workbook.

    python scripts/seed_friends.py --xlsx "~/Downloads/Baptisms (MLC).xlsx" \
        --base http://localhost:8788 --secret <FRIENDS_SYNC_SECRET>

Reads the per-zone tabs, builds the same payload the Apps Script bridge will
send, and POSTs it to /api/friends/sync. After this, the bridge keeps it fresh.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import urllib.request
from datetime import date, timedelta

import openpyxl

ZONE_TABS = [
    "Alexandria", "Annandale", "Bull Run", "Langley", "Loudoun",
    "Manassas", "McLean", "Oakton", "Potomac", "Woodbridge", "Bella Vista North",
]
HEADER_MATCH = "name (first and last)"
FIELD_BY_HEADER = {
    "name (first and last)": "name",
    "baptism date (mm/dd/yy)": "baptismDate",
    "baptism date": "baptismDate",
    "address of baptism": "baptismAddress",
    "time of baptism": "baptismTime",
    "attended church (y/n)": "attendedChurch2x",
    "baptism calendar (y/n)": "onBaptismCalendar",
    "ward name": "ward",
    "stake": "stake",
    "missionary names (last name + last name)": "missionaries",
    "missionary names": "missionaries",
    "completed baptism": "baptizedConfirmed",
}


def most_recent_monday() -> str:
    d = date.today()
    back = 7 if d.weekday() == 6 else d.weekday() + 1  # weekday(): Mon=0, Sun=6
    # back to the most recent past Sunday, then its Monday
    sunday = d - timedelta(days=(d.weekday() + 1) % 7 or 7)
    monday = sunday - timedelta(days=6)
    return monday.isoformat()


def iso_date(v) -> str:
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    s = str(v or "").strip()
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$", s)
    if m:
        yy = ("20" + m.group(3)) if len(m.group(3)) == 2 else m.group(3)
        return f"{yy}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return s


def truthy(v) -> bool:
    return bool(re.match(r"^(y|yes|true|1)", str(v or "").strip(), re.I))


def collect(xlsx: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    out: list[dict] = []
    for tab in ZONE_TABS:
        if tab not in wb.sheetnames:
            continue
        rows = list(wb[tab].iter_rows(values_only=True))
        hdr = next(
            (i for i, r in enumerate(rows)
             if any(str(c or "").strip().lower() == HEADER_MATCH for c in r)),
            None,
        )
        if hdr is None:
            continue
        cols = {}
        for c, label in enumerate(rows[hdr]):
            key = str(label or "").strip().lower()
            if key in FIELD_BY_HEADER:
                cols[FIELD_BY_HEADER[key]] = c
        if "name" not in cols:
            continue
        for r in rows[hdr + 1:]:
            name = str(r[cols["name"]] or "").strip() if cols["name"] < len(r) else ""
            if not name:
                continue
            rec = {"zone": tab, "name": name}
            for field, c in cols.items():
                if field == "name" or c >= len(r):
                    continue
                raw = r[c]
                if field == "baptismDate":
                    rec["baptismDate"] = iso_date(raw) if raw else ""
                elif field in ("attendedChurch2x", "onBaptismCalendar", "baptizedConfirmed"):
                    rec[field] = truthy(raw)
                else:
                    rec[field] = "" if raw is None else str(raw).strip()
            out.append(rec)
    wb.close()
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--base", default="http://localhost:8788")
    ap.add_argument("--secret", default=os.environ.get("FRIENDS_SYNC_SECRET", ""))
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--sql", metavar="FILE", help="write INSERT statements here instead of POSTing")
    args = ap.parse_args()

    rows = collect(os.path.expanduser(args.xlsx))
    payload = {"weekStart": most_recent_monday(), "rows": rows}
    on_date = sum(1 for r in rows if r.get("baptismDate") and not r.get("baptizedConfirmed"))
    print(f"{len(rows)} rows  ({on_date} on date, {len(rows) - on_date} baptized)  week={payload['weekStart']}")

    if args.dry_run:
        print(json.dumps(rows[:3], indent=2))
        return 0

    if args.sql:
        import uuid

        def q(v):
            if v is None or v == "":
                return "NULL"
            return "'" + str(v).replace("'", "''") + "'"

        lines = ["DELETE FROM friend WHERE source = 'sheet';"]
        now = "2026-09-01T00:00:00Z"
        for r in rows:
            key = f"{r.get('zone','')}|{r.get('ward','')}|{r['name']}".lower()
            lines.append(
                "INSERT INTO friend (id,name,zone,ward,stake,missionaries,baptism_date,baptism_time,"
                "baptism_address,attended_church_2x,on_baptism_calendar,baptized_confirmed,dropped,"
                "active,source,sync_key,created_at,created_by,updated_at,updated_by) VALUES ("
                f"{q(str(uuid.uuid4()))},{q(r['name'])},{q(r.get('zone'))},{q(r.get('ward'))},"
                f"{q(r.get('stake'))},{q(r.get('missionaries'))},{q(r.get('baptismDate'))},"
                f"{q(r.get('baptismTime'))},{q(r.get('baptismAddress'))},"
                f"{1 if r.get('attendedChurch2x') else 0},{1 if r.get('onBaptismCalendar') else 0},"
                f"{1 if r.get('baptizedConfirmed') else 0},0,1,'sheet',{q(key)},{q(now)},'seed',{q(now)},'seed');"
            )
        lines.append(
            f"INSERT INTO friend_sync (at,rows_in,upserted,deactivated,warnings) "
            f"VALUES ({q(now)},{len(rows)},{len(rows)},0,NULL);"
        )
        with open(args.sql, "w", encoding="utf-8") as fh:
            fh.write("\n".join(lines) + "\n")
        print(f"wrote {len(lines)} statements to {args.sql}")
        return 0
    if not args.secret:
        raise SystemExit("need --secret or FRIENDS_SYNC_SECRET")

    req = urllib.request.Request(
        args.base.rstrip("/") + "/api/friends/sync",
        data=json.dumps(payload).encode(),
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {args.secret}"},
        method="POST",
    )
    with urllib.request.urlopen(req) as resp:
        print(resp.status, resp.read().decode())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
